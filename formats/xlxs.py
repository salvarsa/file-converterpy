import io
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image, Spacer
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl.utils import get_column_letter
from reportlab.lib.units import cm
    
def get_column_widths(sheet, max_width):
    
    column_widths = []
    for col in sheet.columns:
        width = sheet.column_dimensions[col[0].column_letter].width
        if width is None:
            width = 10
        column_widths.append(width * 8)
    
    total_width = sum(column_widths)
    if total_width > max_width:
        scale = max_width / total_width
        column_widths = [width * scale for width in column_widths]
    
    return column_widths

def get_row_heights(sheet, max_height=50):
    row_heights = []
    for row in sheet.iter_rows():
        height = sheet.row_dimensions[row[0].row].height
        if height is None:
            height = 15
        if height > max_height: 
            height = max_height
        row_heights.append(height)
        
    return row_heights

# Función para procesar celdas y ajustarlas a tamaños de hoja
def process_cell(cell, centered_style, max_characters=100):
    if cell.data_type == 'i':  # Image
        img = Image(io.BytesIO(cell.value))
        img.drawHeight = 50  
        img.drawWidth = 50 
        return img
    elif cell.value is not None:
        cell_value = str(cell.value)
        if len(cell_value) > max_characters: 
            cell_value = cell_value[:max_characters] + '...'  # Truncar contenido largo
        return Paragraph(cell_value, centered_style)
    return ''


def convert_xlsx_to_pdf(input_path: str, output_path: str) -> None:
    wb = load_workbook(input_path)
    sheet = wb.active

    sheet_width = 0
    for column in sheet.columns:
        col_letter = get_column_letter(column[0].column)
        col_width = sheet.column_dimensions[col_letter].width
        if col_width is None:
            col_width = 8.43  # Ancho predeterminado en Excel (8.43 caracteres)
        sheet_width += col_width
    
    sheet_height = 0
    for row in sheet.rows:
        row_height = sheet.row_dimensions[row[0].row].height
        if row_height is None:
            row_height = 15  
        sheet_height += row_height

    # Convertir las dimensiones a puntos (1 punto = 1/72 pulgadas)
    page_width = sheet_width * 0.35  # Aproximación de unidades de Excel a cm
    page_height = sheet_height * 0.035  # Aproximación de unidades de Excel a cm

    margin_left = 0.2 * cm
    margin_right = 0.2 * cm
    margin_top = 0.5 * cm
    margin_bottom = 0.5 * cm

    pdf = SimpleDocTemplate(
        output_path,
        pagesize=(page_width * cm, page_height * cm),
        leftMargin=margin_left,
        rightMargin=margin_right,
        topMargin=margin_top,
        bottomMargin=margin_bottom
    )

    data = []
    styles = getSampleStyleSheet()
    centered_style = ParagraphStyle('centered', parent=styles['Normal'], alignment=TA_CENTER, wordWrap='CJK')
    
    for row in sheet.iter_rows():
        processed_row = []
        for cell in row:
            processed_cell = process_cell(cell, centered_style)
            processed_row.append(processed_cell)
        data.append(processed_row)

    column_widths = get_column_widths(sheet, page_width * cm - (margin_left + margin_right))

    table = Table(data, colWidths=column_widths, repeatRows=1)

    style = TableStyle([
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ])

    # Combinar celdas
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        style.add('SPAN', (min_col - 1, min_row - 1), (max_col - 1, max_row - 1))

    table.setStyle(style)
    elements = [table]

    try:
        pdf.build(elements)
    except Exception as e:
        raise ValueError(f"CONVERTION_ERROR: {e}")